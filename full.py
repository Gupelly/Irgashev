import re
import csv
from datetime import datetime
from prettytable import PrettyTable, ALL
from statistics import mean
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit
# Hello
dict_bool = {"False": "Нет", "True": "Да"}
dict_experience = {"noExperience": "Нет опыта",
                    "between1And3": "От 1 года до 3 лет",
                    "between3And6": "От 3 до 6 лет",
                    "moreThan6": "Более 6 лет"}
dict_currency = {"AZN": "Манаты",
                    "BYR": "Белорусские рубли",
                    "EUR": "Евро",
                    "GEL": "Грузинский лари",
                    "KGS": "Киргизский сом",
                    "KZT": "Тенге",
                    "RUR": "Рубли",
                    "UAH": "Гривны",
                    "USD": "Доллары",
                    "UZS": "Узбекский сум"}
dict_parameter = {'Название': lambda x, parameter: x.name == parameter,
                  'Описание': lambda x, parameter: x.descr == parameter,
                  'Навыки': lambda x, parameter: find_skills(x.skills, parameter),
                  'Опыт работы': lambda x, parameter: dict_experience[x.exp] == parameter,
                  'Премиум-вакансия': lambda x, parameter: dict_bool[x.premium] == parameter,
                  'Компания': lambda x, parameter: x.employer == parameter,
                  'Идентификатор валюты оклада': lambda x, parameter: dict_currency[x.salary.salary_currency] == parameter,
                  'Оклад': lambda x, parameter: x.salary.salary_from <= int(parameter) <= x.salary.salary_to,
                  'Название региона': lambda x, parameter: x.area == parameter,
                  'Дата публикации вакансии': lambda x, parameter:
                  datetime.strptime(x.published_at, '%Y-%m-%dT%H:%M:%S%z').strftime('%d.%m.%Y') == parameter}
dict_sort = {'Название': lambda x: x.name,
                  'Описание': lambda x: x.descr,
                  'Навыки': lambda x: x.skills.count('\n'),
                  'Опыт работы': lambda x: list(dict_experience.keys()).index(x.exp),
                  'Премиум-вакансия': lambda x: dict_bool[x.premium],
                  'Компания': lambda x: x.employer,
                  'Оклад': lambda x: x.salary.average_salary,
                  'Название региона': lambda x: x.area,
                  'Дата публикации вакансии': lambda x: x.published_at}

#Вакансии


def find_skills(x, parameter):
    """
    Проверяет, что навыки вакансии содержат в себе все наваыки, указанные пользователем.

    Args:
        x (str): Все навыки, указанные в вакансии
        parameter (str) : Навыки, введенные пользователем

    Returns:
        bool: Содержат ли навыки вакансии в себе все наваыки, указанные пользователем
    """
    parameter_skills = parameter.split(', ')
    all_skills = x.split()
    for skill in parameter_skills:
        if skill not in all_skills:
            return False
    return True


class DataSet:
    """
    Класс для представления данных csv файла

    Attributes:
        file_name (list): Список заголовков столбцов
        vacancies (list): Список объектов Vacancy
    """

    def __init__(self, file_name):
        """
        Инициализирует объект DataSet, по имени файла получает заголовки столбцов и вакансии без html тегов

        Args:
            file_name (str): Имя файла, введного пользователем
        """
        self.error = False
        reader = self.csv_reader(file_name)
        if reader is None:
            self.error = True
            return
        self.file_name = reader[0]
        self.vacancies = self.csv_filer(reader[1])

    @staticmethod
    def csv_reader(file_name):
        """
        Читает файл, получает заголовки столбцов и вакансии, содуржащие полную информацию

        Args:
            file_name (str) : Имя файла, введного пользователем

        Returns:
            tuple: Список заголовков столбцов, список вакансий
        """
        file_csv = open(file_name, encoding='utf_8_sig')
        reader_csv = csv.reader(file_csv)
        list_data = list(filter(lambda x: '' not in x, reader_csv))
        if len(list_data) == 0:
            return print('Пустой файл')
        if len(list_data[1:]) == 0:
            return print('Нет данных')
        return list_data[0], list_data[1:]

    @staticmethod
    def csv_filer(reader):
        """
        Args:
            reader (str): Список вакансий

        Returns:
            list: Список объектов Vacancy
        """
        vacancies = []
        for line in reader:
            vacancy = []
            for i in range(0, len(line)):
                value = re.sub(re.compile(r"<[^>]*>"), "", line[i])
                value = " ".join(value.split()) if i != 2 else " ".join(value.split(' '))
                vacancy.append(value)
            vacancies.append(Vacancy(vacancy))
        return vacancies


class Salary:
    """
    Класс для предсталения зарплаты

    Attributes:
        __currency_to_rub (dict): Неизменяемый словарь курсов валют
        salary_from (int): Нижняя граница зарплаты
        salary_to (int): Верхняя граница зарплаты
        salary_gross (str): Указывает, включен ли в зарплату вычет налогов
        salary_currency (str): Указывает на курс валюты
        average_salary (float): Средняя зарплата
    """
    def __init__(self, data):
        """
        Инициализирует объект Salary

        Args:
            data (list): Список, состоит из всех значений, необходимых для инициализации объекта Salary
        """
        self.__currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76, "KZT": 0.13,
                       "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}
        self.salary_from = int(float(data[0]))
        self.salary_to = int(float(data[1]))
        self.salary_gross = data[2]
        self.salary_currency = data[3]
        self.average_salary = (self.salary_to + self.salary_from) / 2 * self.__currency_to_rub[data[3]]

    def to_string(self):
        """
        Переводит объект Salary в строку, которая записывается в таблицу

        Returns
            str: Строчная запись объекта Salary
        """
        num_to_str = lambda x: format(x, ",").replace(",", " ")
        return f'{num_to_str(self.salary_from)} - {num_to_str(self.salary_to)} ({dict_currency[self.salary_currency]}) ' \
               f'({"Без вычета налогов" if self.salary_gross == "True" else "С вычетом налогов"})'


class Vacancy:
    """
    Класс для предсталения вакансий

    Attributes:
        name (str): Название специальности
        descr (str): Описание специальности
        skills (str): Требуемые навыки
        exp (str): Требуемый опыт работы
        premium (str): Наличие премиум вакансии
        employer (str): Название компании
        salary (Salary): Зарплата
        area (str): Город
        published_at (str): Дата и время публикации
    """
    def __init__(self, data):
        """
        Инициализирует объект Vacancy

        Args:
            data (list): Список, состоит из всех значений, необходимых для инициализации объекта Vacancy
        """
        self.name = data[0]
        self.descr = data[1]
        self.skills = data[2]
        self.exp = data[3]
        self.premium = data[4]
        self.employer = data[5]
        self.salary = Salary(data[6:10])
        self.area = data[10]
        self.published_at = data[11]

    def translate_vacancy(self):
        """
        Переводит объект Salary в строку, каждый атрибут переводится по собственному правилу

        Returns
            list: Отформатированные атрибуты объекта Vacancy в виде списка
        """
        cut_text = lambda x: x[:100] + '...' if len(x) > 100 else x
        return [self.name, cut_text(self.descr), cut_text(self.skills), dict_experience[self.exp],
                dict_bool[self.premium], self.employer, self.salary.to_string(), self.area,
                datetime.strptime(self.published_at, '%Y-%m-%dT%H:%M:%S%z').strftime('%d.%m.%Y')]


class InputConnect:
    """
    Класс, который получает все данные, введеные пользователем и печатает таблицу

    Attributes:
        __valid_params (list): Неизменяемый список корректных столбцов
        error_message (str): Значение, которое напечатается, если какой-то из введеных параметров некорректен
        filter_param (list or str): list[0] - Фильтруемый столбец, list[1] - Параметр фильтрации, '' если параметр отсутствует
        sort_param (str): Параметр фильтрации, '' если параметр отсутствует
        is_reverse: (str): Параметр порядка последовательности, введенный пользователем
        borders (list): borders[0] - нижняя граница вывода, borders[1] - верхняя граница выводы
        columns (list): Требуемые столбцы
    """
    def __init__(self):
        self.__valid_params = ['Название', 'Описание', 'Навыки', 'Опыт работы', 'Премиум-вакансия', 'Компания', 'Оклад',
        'Название региона', 'Дата публикации вакансии']
        self.error_message = ''
        self.filter_param = self.format_filter_param(input('Введите параметр фильтрации: '))
        self.sort_param = self.format_sort_param(input('Введите параметр сортировки: '))
        self.is_reverse = self.get_bool(input('Обратный порядок сортировки (Да / Нет): '))
        self.borders = [int(x) - 1 for x in input('Введите диапазон вывода: ').split()]
        self.columns = input('Введите требуемые столбцы: ').split(", ")

    def format_filter_param(self, param):
        """
        Проверяет параметр фильтрации на корректность, если он некорректен, меняет значени атрибута error_message

        Args:
            param (str): Параметр фильтрации, введенный пользователем

        Returns:
            list or str: list[0] - Фильтруемый столбец, list[1] - Параметр фильтрации, '' если параметр отсутствует
        """
        if param == '':
            return param
        item = param.split(':')
        if len(item) != 2:
            self.error_message = 'Формат ввода некорректен'
        elif item[0] not in self.__valid_params and item[0] != 'Идентификатор валюты оклада':
            self.error_message = 'Параметр поиска некорректен'
        else:
            item[1] = item[1].lstrip()
        return item

    def format_sort_param(self, param):
        """
        Проверяет параметр сортировки на корректность, если он некорректен, меняет значени атрибута error_message

        Args:
            param (str): Параметр сортировки, введенный пользователем

        Returns:
            string: Параметр фильтрации, '' если параметр отсутствует
        """
        if param != '' and param not in self.__valid_params:
            self.error_message = 'Параметр сортировки некорректен'
        return param

    def get_bool(self, boolen):
        """
        Проверяет параметр порядка последовательности на корректность, если он некорректен, меняет значени атрибута error_message

        Args:
            param (str): Параметр порядка последовательности, введенный пользователем

        Returns:
            bool: 'Да' - True, 'Нет' или '' - False
        """
        if boolen == 'Да':
            return True
        if boolen == 'Нет' or boolen == '':
            return False
        self.error_message = 'Порядок сортировки задан некорректно'

    def print_table(self, vacancies):
        """
        Печатает таблицу

        Args:
            vacancies (list): Список объектов Vacancy

        Returns:
            Выход из списка, если vacancies пустой
        """
        if len(self.filter_param) != 0:
            vacancies = list(filter(lambda x: dict_parameter[self.filter_param[0]](x, self.filter_param[1]), vacancies))
        if self.sort_param != '':
            vacancies.sort(reverse=self.is_reverse, key=dict_sort[self.sort_param])
        if len(vacancies) == 0:
            print('Ничего не найдено')
            return
        vacancies_arr = [x.translate_vacancy() for x in vacancies]

        table = PrettyTable()
        table.field_names = ["№"] + self.__valid_params
        table.align = 'l'
        table.max_width = 20
        table.hrules = ALL
        length = len(vacancies_arr)
        for i in range(0, length):
            table.add_row([i + 1] + vacancies_arr[i])
        print(table.get_string(start=0 if len(self.borders) == 0 else self.borders[0],
                               end=len(vacancies) if len(self.borders) != 2 else self.borders[1],
                               fields=['№'] + (self.columns if self.columns[0] != '' else self.__valid_params)))

# Статистика


class SalaryDict:
    """
    Класс для хранения информация о статистике по зарплате

    Attributes:
        salary_dict (dict): Словарь, key - год или город (str), value - список зарплат (list)
        __average_salary_dict (dict): Неизменяемый словарь, key - год или город (str), value - средняя зарплата (int)
    """
    def __init__(self):
        self.salary_dict = {}
        self.__average_salary_dict = {}

    def add_salary(self, key, salary):
        """
        Добавляет зарплату в список зарплат в salary_dict по ключю

        Args:
            key (str): ключ salary_dict, год или город
            salary (str): зарплата

        Returns:
            None
        """
        if self.salary_dict.get(key) is None:
            self.salary_dict[key] = []
        return self.salary_dict[key].append(salary)

    def get_average_salary(self):
        """
        Находит среднюю зарплату по каждому ключу в salary_dict и добавляет ее в __average_salary_dict по тому же ключу

        Returns:
            dict: __average_salary_dict
        """
        for key, value in self.salary_dict.items():
            self.__average_salary_dict[key] = int(mean(value))
        return self.__average_salary_dict

    def top_salary(self, big_cities):
        """
        Находит до 10 самых крупных средних зарплат по городам, которые есть в big_cities

        Args:
            big_cities (list): Список городов число вакансий в которых больше 1% от общего числа вакансий

        Returns:
            dict: Словарь отсортированный по средним зарплатам для городов, длинной до 10 элементов
        """
        self.get_average_salary()
        sorted_dict = dict(sorted(self.__average_salary_dict.items(), key=lambda x: x[1], reverse=True))
        big_salary_dict = {}
        for key, value in sorted_dict.items():
            if key in big_cities:
                big_salary_dict[key] = value
        return {x: big_salary_dict[x] for x in list(big_salary_dict)[:10]}


class CountDict:
    """
    Класс для хранения информация о статистике по числу вакансий

    Attributes:
        length (int): Общее число всех вакансий
        count_dict (dict): Словарь, key - год или город (str), value - число вакансий
        big_cities (list): Список городов число вакансий в которых больше 1% от общего числа вакансий
        top_proportion_dict (dict): Словарь отсортированный по числу вакансий для городов, длинной до 10 элементов
    """
    def __init__(self):
        self.length = 0
        self.count_dict = {}
        self.big_cities = []
        self.top_proportion_dict = {}

    def add(self, key):
        """
        Увеличить число вакансий в count_dict на 1 по ключу

        Args:
            key (str): ключ count_dict год или город

        Returns:
            None
        """
        if self.count_dict.get(key) is None:
            self.count_dict[key] = 0
        self.count_dict[key] += 1
        self.length += 1
        return

    def get_proportion(self):
        """
        Словарь отсортированный по числу вакансий для городов, длинной до 10 элементов

        Returns:
            None
        """
        proportion_dict = {}
        for key, value in self.count_dict.items():
            proportion = value / self.length
            if proportion >= 0.01:
                self.big_cities.append(key)
                proportion_dict[key] = round(proportion, 4)
        sorted_dict = dict(sorted(proportion_dict.items(), key=lambda x: x[1], reverse=True))
        self.top_proportion_dict = {x: sorted_dict[x] for x in list(sorted_dict)[:10]}
        return


class VacancyForStatistics:
    """
    Класс представляет информацию о вакансиях, необходимую для статистики

    Attributes:
        __dict_currency (dict): Неизменяемый словарь, представляет курс валют
        job (str): Название профессии
        salary (int): Средняя зарплата, с учетом курса валюты
        city (str): Город
        year (str): Год
    """
    def __init__(self, data):
        """
        Args:
            data: Список, состоит из всех значений, необходимых для инициализации объекта VacancyForStatistics
        """
        if len(data) != 6:
            data = [data[0], data[6], data[7], data[9], data[10], data[11]]
        self.__dict_currency = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76, "KZT": 0.13,
                   "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}
        self.job = data[0]
        self.salary = (float(data[1]) + float(data[2])) / 2 * self.__dict_currency[data[3]]
        self.city = data[4]
        self.year = int(datetime.strptime(data[5], '%Y-%m-%dT%H:%M:%S%z').strftime('%Y'))


class Result:
    """
    Класс, который получает все данные, введеные пользователем и форматирует их

    Attributes:
        job (str): Название профессии
        salary_year (SalaryDict): зарплаты по годам
        count_year (CountDict): число вакансий погодам
        job_salary_year (SalaryDict): зарплаты по годам для выбранной специальности
        job_count_year (CountDict): число вакансий погодам для выбранной специальности
        job_salary_city (SalaryDict): зарплаты по городам для выбранной специальности
        job_count_city (CountDict): число вакансий погородам для выбранной специальности
    """
    def __init__(self, job):
        """
        Args:
            job: Название специальности
        """
        self.job = job
        self.salary_year = SalaryDict()
        self.count_year = CountDict()
        self.job_salary_year = SalaryDict()
        self.job_count_year = CountDict()
        self.job_salary_city = SalaryDict()
        self.job_count_city = CountDict()

    def get_data(self, vacancies):
        """
        Получает необходимые данные из списка объектов Vacancy по запросам пользователя и добавляет их в
        соответствующие атрибуты

        Args:
            vacancies (list): список объектов Vacancy

        Returns:
            None
        """
        for vacancy in vacancies:
            self.salary_year.add_salary(vacancy.year, vacancy.salary)
            self.count_year.add(vacancy.year)
            self.job_salary_city.add_salary(vacancy.city, vacancy.salary)
            self.job_count_city.add(vacancy.city)
            if self.job in vacancy.job:
                self.job_salary_year.add_salary(vacancy.year, vacancy.salary)
                self.job_count_year.add(vacancy.year)
        if self.job_salary_year.salary_dict == {}:
            self.job_salary_year.salary_dict = {x: [0] for x in self.salary_year.salary_dict.keys()}
        if self.job_count_year.count_dict == {}:
            self.job_count_year.count_dict = {x: 0 for x in self.count_year.count_dict.keys()}
        self.job_count_city.get_proportion()
        return

    def print_result(self):
        """
        Печатает требуемые данные

        Returns:
            None
        """
        print(f'Динамика уровня зарплат по годам: {self.salary_year.get_average_salary()}')
        print(f'Динамика количества вакансий по годам: {self.count_year.count_dict}')
        print(f'Динамика уровня зарплат по годам для выбранной профессии: {self.job_salary_year.get_average_salary()}')
        print(f'Динамика количества вакансий по годам для выбранной профессии: {self.job_count_year.count_dict}')
        print(f'Уровень зарплат по городам (в порядке убывания): {self.job_salary_city.top_salary(self.job_count_city.big_cities)}')
        print(f'Доля вакансий по городам (в порядке убывания): {self.job_count_city.top_proportion_dict}')
        return [self.salary_year.get_average_salary(), self.job_salary_year.get_average_salary()], \
               [self.count_year.count_dict, self.job_count_year.count_dict], self.job_salary_city.top_salary(self.job_count_city.big_cities),\
            self.job_count_city.top_proportion_dict

    def get_excel_data(self):
        """
        Возвращает данные, необходимые для создания excel файла

        Returns:
            None
        """
        salary_list = [['Год', 'Средняя зарплата', f'Средняя зарплата - {self.job}', 'Количество вакансий',
                        f'Количество вакансий - {self.job}']]
        for year in self.salary_year.salary_dict:
            salary_list.append([year, self.salary_year.get_average_salary()[year], self.job_salary_year.get_average_salary()[year],
                         self.count_year.count_dict[year], self.job_count_year.count_dict[year]])
        city_list = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]
        city_salary = list(self.job_salary_city.top_salary(self.job_count_city.big_cities).items())
        city_count = list(self.job_count_city.top_proportion_dict.items())
        for i in range(len(city_count)):
            city_list.append([city_salary[i][0], city_salary[i][1], '', city_count[i][0], city_count[i][1]])
        return salary_list, city_list

    def get_png_data(self):
        """
        Возвращает данные, необходимые для создания png файла

        Returns:
            None
        """
        return [self.salary_year.get_average_salary(), self.job_salary_year.get_average_salary()], \
               [self.count_year.count_dict, self.job_count_year.count_dict], self.job_salary_city.top_salary(self.job_count_city.big_cities),\
               self.job_count_city.top_proportion_dict


class ReportExcel:
    """
    Класс, который создает excel таблицу

    Attributes:
        wb (Workbook): excel таблица
    """
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb['Sheet'])

    def create_sheet(self, title, data, percent=False):
        """
        Создание excel таблицы

        Args:
            title (str): Название таблицы
            data (list): Содержание таблицы
            percent (bool): Добавить значение процентов для последнего столбца

        Returns:
            None
        """
        ws = self.wb.create_sheet(title)
        for line in data:
            ws.append(line)

        for i in range(len(list(ws.columns))):
            ws.cell(row=1, column=i+1).font = Font(bold=True)
        for column in ws.columns:
            length = max(len(str(cell.value)) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = length + 2
            for cell in column:
                if str(cell.value) != '':
                    thin = Side(border_style="thin", color="000000")
                    cell.border = Border(left=thin, top=thin, right=thin, bottom=thin)
        if percent:
            for i in range(len(list(ws.rows)) - 1):
                ws.cell(row=i + 2, column=5).number_format = '0.00%'

    def save_wb(self):
        """
        Сохраняет таблицу

        Returns:
            None
        """
        self.wb.save('report.xlsx')


class ReportPng:
    """
        Создание графиков в формате png

        Attributes:
            fig: График
            width (float): Ширина обводки
            num: Положение графика

    """
    def __init__(self):
        self.fig = plt.figure()
        self.width = 0.4
        self.num = 220

    def add_empty_graph(self, title):
        """
        Создает 'пустой' график с названием

        Args:
            title (str): Название графика

        Returns:
            График
        """
        self.num += 1
        ax = self.fig.add_subplot(self.num)
        ax.set_title(title)
        return ax

    def add_graph(self, title, label_list, dict_list):
        """
        Создает график

        Args:
            title (str): Название графика
            label_list (list): Список по годам, городам
            dict_list (dict): Словарь рядов словаря

        Returns:
            None
        """
        ax = self.add_empty_graph(title)
        x_nums = np.arange(len(dict_list[0].keys()))
        x_list = [x_nums - self.width / 2, x_nums + self.width / 2]
        for i in range(len(label_list)):
            ax.bar(x_list[i], dict_list[i].values(), self.width, label=label_list[i])
        ax.set_xticks(x_nums, dict_list[0].keys(), rotation='vertical')
        ax.tick_params(axis='both', labelsize=8)
        ax.legend(fontsize=8)
        ax.grid(True, axis='y')

    def add_turned_graph(self, title, city_salary):
        """
        Создает перевернутый график

        Args:
            title (str): Название графика
            city_salary (dict): Список с данными

        Returns:
            None
        """
        ax = self.add_empty_graph(title)
        x_nums = np.arange(len(city_salary.keys()))
        ax.barh(x_nums, city_salary.values(), self.width * 2)
        keys = ['\n'.join(x.split()) for x in city_salary.keys()]
        keys = ['-\n'.join(x.split('-')) for x in keys]
        ax.set_yticks(x_nums, keys)
        ax.invert_yaxis()
        ax.tick_params(axis='x', labelsize=8)
        ax.tick_params(axis='y', labelsize=6)
        ax.grid(True, axis='x')

    def add_round_graph(self, title, city_count):
        """
        Создает круговую диаграмму

        Args:
            title (str): Название графика
            city_count (dict): Список с данными

        Returns:
            None
        """
        ax = self.add_empty_graph(title)
        city_count = {'Другие': 1 - sum(city_count.values())} | city_count
        ax.pie(city_count.values(), labels=city_count.keys(), textprops={'fontsize': 6})

    @staticmethod
    def print_graph():
        """
        Сохраняет графики

        Returns:
            None
        """
        plt.tight_layout()
        plt.savefig("graph.png")
        plt.show()


def csv_reader(file_name):
    """
    Читает файл, получает заголовки столбцов и вакансии, содуржащие полную информацию

    Args:
        file_name (str) : Имя файла, введного пользователем

    Returns:
        list: список вакансий
    """
    file_csv = open(file_name, encoding='utf_8_sig')
    reader_csv = csv.reader(file_csv)
    list_data = list(filter(lambda x: '' not in x, reader_csv))
    if len(list_data) == 0:
        return print('Пустой файл')
    if len(list_data[1:]) == 0:
        return print('Нет данных')
    return list_data[1:]


choose = input('Выберите формат выходных данных ')
if choose == 'Вакансии':
    file_name = input('Введите название файла: ')
    input_data = InputConnect()
    if input_data.error_message != '':
        print(input_data.error_message)
    else:
        data = DataSet(file_name)
        if not data.error:
            input_data.print_table(data.vacancies)
elif choose == 'Статистика':
    file_name = input('Введите название файла: ')
    job = input('Введите название профессии: ')
    file = open(file_name, encoding='utf_8_sig')
    data = csv_reader(file_name)
    if data is not None:
        vacancies = [VacancyForStatistics(x) for x in data]
        result = Result(job)
        result.get_data(vacancies)
        result.print_result()

        wb = ReportExcel()
        salary_list, city_list = result.get_excel_data()
        wb.create_sheet('Статистика по годам', salary_list)
        wb.create_sheet('Статистика по городам', city_list, True)
        wb.save_wb()

        fig = ReportPng()
        salary_year, salary_count, city_salary, city_count = result.get_png_data()
        fig.add_graph("Уровень зарплат по годам", ['средняя з/п', f'з/п {job}'], salary_year)
        fig.add_graph("Количество вакансий по годам", ['Количество вакансий', f'Количество вакансий {job}'],
                      salary_count)
        fig.add_turned_graph("Уровень зарплат по городам", city_salary)
        fig.add_round_graph("Доля вакансий по городам", city_count)
        fig.print_graph()

        table1 = [x[:2] for x in city_list]
        table2 = [x[3:] for x in city_list]
        for row in table2[1:]:
            row[1] = ("{:.2%}".format(row[1]).replace('.', ','))

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        pdf_template = template.render({'job': job, 'table_big': salary_list, 'table1': table1, 'table2': table2})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": None})
else:
    print('Некорректный формат выходных данных')