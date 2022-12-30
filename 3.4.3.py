import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit


class DataSet:

    def __init__(self, file_name):
        self.file = pd.read_csv(file_name)
        path = r'C:\Users\user\Irgashev\years'
        for f in os.listdir(path):
            os.remove(os.path.join(path, f))
        self.files_by_years = self.get_folders(self.file)

    @staticmethod
    def get_folders(df):
        if df is None:
            return
        df['years'] = df['published_at'].apply(lambda x: int(x[:4]))
        years = df['years'].unique()
        files_by_year = {}
        for year in years:
            data = df[df['years'] == year]
            year_file = data[['name', 'salary_from', 'salary_to', 'salary_currency', 'area_name', 'published_at']]
            files_by_year[year] = year_file
            year_file.to_csv(rf'years\{year}.csv', index=False)
        return files_by_year


class Statistics:
    def __init__(self, job, region, big_file, files_by_years):
        self.job = job
        self.region = region
        self.big_file = big_file
        self.files_by_years = files_by_years
        self.salary_statistic = self.get_salary_statistic()
        self.city_statistic = self.get_city_statistic()

    def one_year_statistic(self, item):
        df = item[1]
        df['salary'] = df[['salary_from', 'salary_to']].mean(axis=1)
        df_job = df[(df['name'].str.contains(self.job)) & (df['area_name'] == self.region)]
        return [item[0], int(df_job['salary'].mean()), len(df_job)]

    def get_salary_statistic(self):
        result = []
        for item in self.files_by_years.items():
            result.append(self.one_year_statistic(item))
        return result

    def get_city_statistic(self):
        total = len(self.big_file)
        self.big_file['salary'] = self.big_file[['salary_from', 'salary_to']].mean(axis=1)
        self.big_file['count'] = self.big_file.groupby('area_name')['area_name'].transform('count')
        df_big = self.big_file[self.big_file['count'] > 0.01 * total].groupby('area_name', as_index=False)

        df_salary_area = df_big['salary'].mean().sort_values(by='salary', ascending=False)
        df_top_salary_area = df_salary_area.head(10)
        top_salary = df_salary_area['salary'].apply(lambda x: int(x))
        salary_by_cities = dict(zip(df_top_salary_area['area_name'], top_salary))

        df_count_area = df_big['count'].mean().sort_values(by='count', ascending=False)
        df_top_count_area = df_count_area.head(10)
        top_cities = df_top_count_area['count'].apply(lambda x: round(x / total, 4))
        count_by_cities = dict(zip(df_top_count_area['area_name'], top_cities))
        return salary_by_cities, count_by_cities

    def get_excel_data(self):
        city_data = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]
        city_salary = list(self.city_statistic[0].items())
        city_count = list(self.city_statistic[1].items())
        for i in range(len(city_salary)):
            city_data.append([city_salary[i][0], city_salary[i][1], '', city_count[i][0], city_count[i][1]])
        return [['Год', f'Средняя зарплата - {self.region}, {self.job}', f'Количество вакансий - {self.region}, {self.job}']] + self.salary_statistic, city_data

    def get_png_data(self):
        return {x[0]: x[1] for x in self.salary_statistic}, {x[0]: x[2] for x in self.salary_statistic}, \
               self.city_statistic[0], self.city_statistic[1]


class ReportExcel:
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb['Sheet'])

    def create_sheet(self, title, data, percent=False):
        ws = self.wb.create_sheet(title)
        for line in data:
            ws.append(line)

        for i in range(len(list(ws.columns))):
            ws.cell(row=1, column=i + 1).font = Font(bold=True)
        for column in ws.columns:
            length = max(len(str(cell.value)) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = length + 2
            for cell in column:
                if str(cell.value) != '':
                    thin = Side(border_style="thin", color="000000")
                    cell.border = Border(left=thin, top=thin, right=thin, bottom=thin)
        if percent:
            for i in range(len(list(ws.rows)) - 1):
                ws.cell(row=i + 2, column=5).number_format = '0.00%'

    def save_wb(self):
        self.wb.save('3.4.3_report.xlsx')


class ReportPng:
    def __init__(self):
        self.fig = plt.figure()
        self.width = 0.4
        self.num = 220

    def add_empty_graph(self, title):
        self.num += 1
        ax = self.fig.add_subplot(self.num)
        ax.set_title(title)
        return ax

    def add_graph(self, title, label, stat):
        ax = self.add_empty_graph(title)
        ax.bar(stat.keys(), stat.values(), self.width, label=label)
        ax.tick_params(axis='both', labelsize=8)
        ax.legend(fontsize=8)
        ax.grid(True, axis='y')

    def add_turned_graph(self, title, city_salary):
        ax = self.add_empty_graph(title)
        x_nums = np.arange(len(city_salary.keys()))
        ax.barh(x_nums, city_salary.values(), self.width * 2)
        keys = ['\n'.join(x.split()) for x in city_salary.keys()]
        keys = ['-\n'.join(x.split('-')) for x in keys]
        ax.set_yticks(x_nums, keys)
        ax.invert_yaxis()
        ax.tick_params(axis='x', labelsize=8)
        ax.tick_params(axis='y', labelsize=6)
        ax.grid(True, axis='x')

    def add_round_graph(self, title, city_count):
        ax = self.add_empty_graph(title)
        city_count = {'Другие': 1 - sum(city_count.values())} | city_count
        ax.pie(city_count.values(), labels=city_count.keys(), textprops={'fontsize': 6})

    @staticmethod
    def print_graph():
        plt.tight_layout()
        plt.savefig("3.4.3_graph.png")
        plt.show()


data = DataSet('vacancies_by_year.csv')
job = 'Аналитик'
region = 'Москва'
statistic = Statistics(job, region, data.file, data.files_by_years)

wb = ReportExcel()
salary_list, city_list = statistic.get_excel_data()
wb.create_sheet('Статистика по годам', salary_list)
wb.create_sheet('Статистика по городам', city_list, True)
wb.save_wb()

fig = ReportPng()
salary_year, salary_count, city_salary, city_count = statistic.get_png_data()
fig.add_turned_graph("Уровень зарплат по городам", city_salary)
fig.add_round_graph("Доля вакансий по городам", city_count)
fig.add_graph("Уровень зарплат по годам", f'з/п {region}, {job}', salary_year)
fig.add_graph("Количество вакансий по годам", f'Количество вакансий {region}, {job}', salary_count)
fig.print_graph()

table1 = [x[:2] for x in city_list]
table2 = [x[3:] for x in city_list]
for row in table2[1:]:
    row[1] = ("{:.2%}".format(row[1]).replace('.', ','))
env = Environment(loader=FileSystemLoader('.'))
template = env.get_template("3.4.3_pdf_template.html")
pdf_template = template.render({'job': job, 'region': region, 'table_big': salary_list, 'table1': table1, 'table2': table2})
config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
pdfkit.from_string(pdf_template, '3.4.3_report.pdf', configuration=config, options={"enable-local-file-access": None})
