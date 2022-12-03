import csv
from datetime import datetime
from statistics import mean
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit

# Hiiii
class SalaryDict:
    def __init__(self):
        self.salary_dict = {}
        self.__average_salary_dict = {}

    def add_salary(self, key, salary):
        if self.salary_dict.get(key) is None:
            self.salary_dict[key] = []
        return self.salary_dict[key].append(salary)

    def get_average_salary(self):
        for key, value in self.salary_dict.items():
            self.__average_salary_dict[key] = int(mean(value))
        return self.__average_salary_dict

    def top_salary(self, big_cities):
        self.get_average_salary()
        sorted_dict = dict(sorted(self.__average_salary_dict.items(), key=lambda x: x[1], reverse=True))
        big_salary_dict = {}
        for key, value in sorted_dict.items():
            if key in big_cities:
                big_salary_dict[key] = value
        return {x: big_salary_dict[x] for x in list(big_salary_dict)[:10]}


class CountDict:
    def __init__(self):
        self.length = 0
        self.count_dict = {}
        self.big_cities = []
        self.top_proportion_dict = {}

    def add(self, key):
        if self.count_dict.get(key) is None:
            self.count_dict[key] = 0
        self.count_dict[key] += 1
        self.length += 1
        return

    def get_proportion(self):
        proportion_dict = {}
        for key, value in self.count_dict.items():
            proportion = value / self.length
            if proportion >= 0.01:
                self.big_cities.append(key)
                proportion_dict[key] = round(proportion, 4)
        sorted_dict = dict(sorted(proportion_dict.items(), key=lambda x: x[1], reverse=True))
        self.top_proportion_dict = {x: sorted_dict[x] for x in list(sorted_dict)[:10]}
        return


class Vacancy:
    def __init__(self, data):
        if len(data) != 6:
            data = [data[0], data[6], data[7], data[9], data[10], data[11]]
        self.__dict_currency = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76, "KZT": 0.13,
                   "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}
        self.job = data[0]
        self.salary = (float(data[1]) + float(data[2])) / 2 * self.__dict_currency[data[3]]
        self.city = data[4]
        self.year = int(datetime.strptime(data[5], '%Y-%m-%dT%H:%M:%S%z').strftime('%Y'))


class Result:
    def __init__(self, job):
        self.job = job
        self.salary_year = SalaryDict()
        self.count_year = CountDict()
        self.job_salary_year = SalaryDict()
        self.job_count_year = CountDict()
        self.job_salary_city = SalaryDict()
        self.job_count_city = CountDict()

    def get_data(self, vacancies):
        for vacancy in vacancies:
            self.salary_year.add_salary(vacancy.year, vacancy.salary)
            self.count_year.add(vacancy.year)
            self.job_salary_city.add_salary(vacancy.city, vacancy.salary)
            self.job_count_city.add(vacancy.city)
            if self.job in vacancy.job:
                self.job_salary_year.add_salary(vacancy.year, vacancy.salary)
                self.job_count_year.add(vacancy.year)
        if self.job_salary_year.salary_dict == {}:
            self.job_salary_year.salary_dict = {x: [0] for x in self.salary_year.salary_dict.keys()}
        if self.job_count_year.count_dict == {}:
            self.job_count_year.count_dict = {x: 0 for x in self.count_year.count_dict.keys()}
        self.job_count_city.get_proportion()
        return

    def print_result(self):
        print(f'Динамика уровня зарплат по годам: {self.salary_year.get_average_salary()}')
        print(f'Динамика количества вакансий по годам: {self.count_year.count_dict}')
        print(f'Динамика уровня зарплат по годам для выбранной профессии: {self.job_salary_year.get_average_salary()}')
        print(f'Динамика количества вакансий по годам для выбранной профессии: {self.job_count_year.count_dict}')
        print(f'Уровень зарплат по городам (в порядке убывания): {self.job_salary_city.top_salary(self.job_count_city.big_cities)}')
        print(f'Доля вакансий по городам (в порядке убывания): {self.job_count_city.top_proportion_dict}')
        return [self.salary_year.get_average_salary(), self.job_salary_year.get_average_salary()], \
               [self.count_year.count_dict, self.job_count_year.count_dict], self.job_salary_city.top_salary(self.job_count_city.big_cities),\
            self.job_count_city.top_proportion_dict

    def get_excel_data(self):
        salary_list = [['Год', 'Средняя зарплата', f'Средняя зарплата - {self.job}', 'Количество вакансий',
                        f'Количество вакансий - {self.job}']]
        for year in self.salary_year.salary_dict:
            salary_list.append([year, self.salary_year.get_average_salary()[year], self.job_salary_year.get_average_salary()[year],
                         self.count_year.count_dict[year], self.job_count_year.count_dict[year]])
        city_list = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]
        city_salary = list(self.job_salary_city.top_salary(self.job_count_city.big_cities).items())
        city_count = list(self.job_count_city.top_proportion_dict.items())
        for i in range(len(city_count)):
            city_list.append([city_salary[i][0], city_salary[i][1], '', city_count[i][0], city_count[i][1]])
        return salary_list, city_list

    def get_png_data(self):
        return [self.salary_year.get_average_salary(), self.job_salary_year.get_average_salary()], \
               [self.count_year.count_dict, self.job_count_year.count_dict], self.job_salary_city.top_salary(self.job_count_city.big_cities),\
               self.job_count_city.top_proportion_dict


class ReportExcel:
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb['Sheet'])

    def create_sheet(self, title, data, percent=False):
        ws = self.wb.create_sheet(title)
        for line in data:
            ws.append(line)

        for i in range(len(list(ws.columns))):
            ws.cell(row=1, column=i+1).font = Font(bold=True)
        for column in ws.columns:
            length = max(len(str(cell.value)) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = length + 2
            for cell in column:
                if str(cell.value) != '':
                    thin = Side(border_style="thin", color="000000")
                    cell.border = Border(left=thin, top=thin, right=thin, bottom=thin)
        if percent:
            for i in range(len(list(ws.rows)) - 1):
                ws.cell(row=i + 2, column=5).number_format = '0.00%'

    def save_wb(self):
        self.wb.save('report.xlsx')


class ReportPng:
    def __init__(self):
        self.fig = plt.figure()
        self.width = 0.4
        self.num = 220

    def add_empty_graph(self, title):
        self.num += 1
        ax = self.fig.add_subplot(self.num)
        ax.set_title(title)
        return ax

    def add_graph(self, title, label_list, dict_list):
        ax = self.add_empty_graph(title)
        x_nums = np.arange(len(dict_list[0].keys()))
        x_list = [x_nums - self.width / 2, x_nums + self.width / 2]
        for i in range(len(label_list)):
            ax.bar(x_list[i], dict_list[i].values(), self.width, label=label_list[i])
        ax.set_xticks(x_nums, dict_list[0].keys(), rotation='vertical')
        ax.tick_params(axis='both', labelsize=8)
        ax.legend(fontsize=8)
        ax.grid(True, axis='y')

    def add_turned_graph(self, title, city_salary):
        ax = self.add_empty_graph(title)
        x_nums = np.arange(len(city_salary.keys()))
        ax.barh(x_nums, city_salary.values(), self.width * 2)
        keys = ['\n'.join(x.split()) for x in city_salary.keys()]
        keys = ['-\n'.join(x.split('-')) for x in keys]
        ax.set_yticks(x_nums, keys)
        ax.invert_yaxis()
        ax.tick_params(axis='x', labelsize=8)
        ax.tick_params(axis='y', labelsize=6)
        ax.grid(True, axis='x')

    def add_round_graph(self, title, city_count):
        ax = self.add_empty_graph(title)
        city_count = {'Другие': 1 - sum(city_count.values())} | city_count
        ax.pie(city_count.values(), labels=city_count.keys(), textprops={'fontsize': 6})

    @staticmethod
    def print_graph():
        plt.tight_layout()
        plt.savefig("graph.png")
        plt.show()


def csv_reader(file_name):
    file_csv = open(file_name, encoding='utf_8_sig')
    reader_csv = csv.reader(file_csv)
    list_data = list(filter(lambda x: '' not in x, reader_csv))
    if len(list_data) == 0:
        return print('Пустой файл')
    if len(list_data[1:]) == 0:
        return print('Нет данных')
    return list_data[1:]


file_name = 'vacancies_by_year.csv'
job = input('Введите название профессии: ')
file = open(file_name, encoding='utf_8_sig')
data = csv_reader(file_name)
if data is not None:
    vacancies = [Vacancy(x) for x in data]
    result = Result(job)
    result.get_data(vacancies)
    result.print_result()

    wb = ReportExcel()
    salary_list, city_list = result.get_excel_data()
    wb.create_sheet('Статистика по годам', salary_list)
    wb.create_sheet('Статистика по городам', city_list, True)
    wb.save_wb()

    fig = ReportPng()
    salary_year, salary_count, city_salary, city_count = result.get_png_data()
    fig.add_graph("Уровень зарплат по годам", ['средняя з/п', f'з/п {job}'], salary_year)
    fig.add_graph("Количество вакансий по годам", ['Количество вакансий', f'Количество вакансий {job}'], salary_count)
    fig.add_turned_graph("Уровень зарплат по городам", city_salary)
    fig.add_round_graph("Доля вакансий по городам", city_count)
    fig.print_graph()

    table1 = [x[:2] for x in city_list]
    table2 = [x[3:] for x in city_list]
    for row in table2[1:]:
        row[1] = ("{:.2%}".format(row[1]).replace('.', ','))

    env = Environment(loader=FileSystemLoader('.'))
    template = env.get_template("pdf_template.html")
    pdf_template = template.render({'job': job, 'table_big': salary_list, 'table1': table1, 'table2': table2})
    config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
    pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": None})
